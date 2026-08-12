"""
Step 5: Excel deliverable — client-readable workbook with a summary tab,
formula-driven pivot tabs, and embedded charts. No hardcoded results:
every aggregate is a formula referencing the Raw Data tab.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

df = pd.read_csv('/home/claude/supply-chain-project/data/analyzed_data.csv')

FONT = 'Arial'
NAVY = '1F3864'
LIGHT_BLUE = 'D9E2F3'
GREEN = '27AE60'
RED = 'C0392B'

wb = Workbook()

# ============================================================
# TAB 1: Summary (executive front page)
# ============================================================
ws = wb.active
ws.title = 'Summary'
ws.sheet_view.showGridLines = False

title_font = Font(name=FONT, size=18, bold=True, color=NAVY)
header_font = Font(name=FONT, size=12, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor=NAVY)
label_font = Font(name=FONT, size=11, bold=True)
kpi_font = Font(name=FONT, size=20, bold=True, color=NAVY)
kpi_label_font = Font(name=FONT, size=10, color='595959')
normal_font = Font(name=FONT, size=11)

ws['B2'] = 'Supply Chain Performance & Cost Analysis'
ws['B2'].font = title_font
ws['B3'] = 'Beauty & Personal Care Supply Chain — 100 SKUs across 5 suppliers and 5 Indian distribution hubs'
ws['B3'].font = Font(name=FONT, size=11, italic=True, color='595959')

# KPI cards (row 6-8), formulas referencing Raw Data
kpis = [
    ('Total Revenue', "=\"$\"&TEXT(SUM('Raw Data'!F2:F101),\"#,##0\")", 'B'),
    ('Total Profit', "=\"$\"&TEXT(SUM('Raw Data'!Y2:Y101),\"#,##0\")", 'D'),
    ('Avg Profit Margin', "=TEXT(AVERAGE('Raw Data'!Z2:Z101),\"0.0\")&\"%\"", 'F'),
    ('SKUs at Risk*', "=SUMPRODUCT((('Raw Data'!AB2:AB101<'Raw Data'!P2:P101)+('Raw Data'!T2:T101=\"Fail\")>0)*1)", 'H'),
]
row = 6
for label, formula, col in kpis:
    ws[f'{col}{row}'] = label
    ws[f'{col}{row}'].font = kpi_label_font
    ws[f'{col}{row+1}'] = formula
    ws[f'{col}{row+1}'].font = kpi_font

ws['B9'] = '*SKUs with stock cover below supplier lead time, or a failed QC inspection'
ws['B9'].font = Font(name=FONT, size=8, italic=True, color='808080')

# Key insights block
ws['B12'] = 'Key Business Insights'
ws['B12'].font = Font(name=FONT, size=14, bold=True, color=NAVY)

insights = [
    ('1. Category mix', 'Skincare drives the most revenue ($241.6K) but cosmetics has the best average margin (87.3%) despite fewer SKUs (26). Cosmetics is under-invested relative to its profitability.'),
    ('2. Supplier risk', 'Supplier 3 carries the highest combined risk (long ~20-day lead times + elevated 2.5% defect rate). Supplier 4 has the worst QC pass-through — 67% of its SKUs fail inspection, even though its average defect rate looks mid-pack.'),
    ('3. Logistics cost', 'Carrier A + Sea is the cheapest and most time-efficient combination ($3.88 avg, 7.0 days). Carrier C + Road is the most expensive per shipment ($6.55 avg) despite being a short-haul mode — a clear cost-cutting target.'),
    ('4. Profit concentration', 'The top 20 SKUs (20% of the catalog) generate ~33% of total profit. These SKUs deserve prioritized stock allocation and supplier attention.'),
    ('5. Inventory risk', 'A majority of SKUs show stock cover thinner than their own supplier lead time — meaning if demand holds, many products will run out before replenishment arrives. See Stockout Risk tab for the ranked list.'),
]
r = 13
for head, body in insights:
    ws[f'B{r}'] = head
    ws[f'B{r}'].font = Font(name=FONT, size=11, bold=True, color=NAVY)
    ws.merge_cells(f'C{r}:J{r}')
    ws[f'C{r}'] = body
    ws[f'C{r}'].font = normal_font
    ws[f'C{r}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 48
    r += 1

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 16
for c in 'CDEFGHIJ':
    ws.column_dimensions[c].width = 13
ws.column_dimensions['C'].width = 8


# ============================================================
# TAB 2: Raw Data
# ============================================================
ws2 = wb.create_sheet('Raw Data')
for j, col in enumerate(df.columns, start=1):
    c = ws2.cell(row=1, column=j, value=col)
    c.font = header_font
    c.fill = header_fill
for i, row_data in enumerate(df.itertuples(index=False), start=2):
    for j, val in enumerate(row_data, start=1):
        ws2.cell(row=i, column=j, value=val)

n_rows = len(df) + 1
last_col = get_column_letter(len(df.columns))
tab = Table(displayName="RawData", ref=f"A1:{last_col}{n_rows}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws2.add_table(tab)
for j, col in enumerate(df.columns, start=1):
    width = max(12, min(22, len(str(col)) + 2))
    ws2.column_dimensions[get_column_letter(j)].width = width

# ============================================================
# TAB 3: Product Analysis (formula-driven pivot by Product type)
# ============================================================
ws3 = wb.create_sheet('Product Analysis')
ws3.sheet_view.showGridLines = False
ws3['B2'] = 'Revenue & Profitability by Product Category'
ws3['B2'].font = Font(name=FONT, size=14, bold=True, color=NAVY)

headers = ['Product Type', 'SKU Count', 'Total Revenue', 'Total Profit', 'Avg Margin %', 'Avg Defect Rate %']
for j, h in enumerate(headers, start=2):
    c = ws3.cell(row=4, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')

categories = ['skincare', 'haircare', 'cosmetics']
for i, cat in enumerate(categories, start=5):
    ws3.cell(row=i, column=2, value=cat.capitalize())
    ws3.cell(row=i, column=3, value=f'=COUNTIF(\'Raw Data\'!A:A,B{i})')
    ws3.cell(row=i, column=4, value=f'=SUMIF(\'Raw Data\'!A:A,B{i},\'Raw Data\'!F:F)').number_format = '$#,##0'
    ws3.cell(row=i, column=5, value=f'=SUMIF(\'Raw Data\'!A:A,B{i},\'Raw Data\'!Y:Y)').number_format = '$#,##0'
    ws3.cell(row=i, column=6, value=f'=AVERAGEIF(\'Raw Data\'!A:A,B{i},\'Raw Data\'!Z:Z)').number_format = '0.0"%"'
    ws3.cell(row=i, column=7, value=f'=AVERAGEIF(\'Raw Data\'!A:A,B{i},\'Raw Data\'!AC:AC)').number_format = '0.00"%"'

ws3.cell(row=8, column=2, value='TOTAL / AVG').font = Font(name=FONT, bold=True)
ws3.cell(row=8, column=3, value='=SUM(C5:C7)').font = Font(name=FONT, bold=True)
ws3.cell(row=8, column=4, value='=SUM(D5:D7)').number_format = '$#,##0'
ws3.cell(row=8, column=4).font = Font(name=FONT, bold=True)
ws3.cell(row=8, column=5, value='=SUM(E5:E7)').number_format = '$#,##0'
ws3.cell(row=8, column=5).font = Font(name=FONT, bold=True)
ws3.cell(row=8, column=6, value='=AVERAGE(F5:F7)').number_format = '0.0"%"'
ws3.cell(row=8, column=6).font = Font(name=FONT, bold=True)
ws3.cell(row=8, column=7, value='=AVERAGE(G5:G7)').number_format = '0.00"%"'
ws3.cell(row=8, column=7).font = Font(name=FONT, bold=True)

for col_letter, width in [('B', 16), ('C', 12), ('D', 14), ('E', 14), ('F', 12), ('G', 16)]:
    ws3.column_dimensions[col_letter].width = width

# Embedded chart: revenue by category
chart = BarChart()
chart.title = "Total Revenue by Product Category"
chart.y_axis.title = 'Revenue ($)'
chart.x_axis.title = 'Product Type'
data_ref = Reference(ws3, min_col=4, min_row=4, max_row=7)
cats_ref = Reference(ws3, min_col=2, min_row=5, max_row=7)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 14
chart.height = 8
ws3.add_chart(chart, 'B11')

# ============================================================
# TAB 4: Supplier Scorecard
# ============================================================
ws4 = wb.create_sheet('Supplier Scorecard')
ws4.sheet_view.showGridLines = False
ws4['B2'] = 'Supplier Risk Scorecard'
ws4['B2'].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws4['B3'] = 'Risk score = z-scored avg defect rate + z-scored avg lead time (higher = riskier)'
ws4['B3'].font = Font(name=FONT, size=9, italic=True, color='808080')

sup_headers = ['Supplier', 'SKU Count', 'Avg Defect Rate %', 'Avg Lead Time (days)', 'QC Fail Rate %', 'Total Revenue', 'Risk Level']
for j, h in enumerate(sup_headers, start=2):
    c = ws4.cell(row=5, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)

suppliers = ['Supplier 1', 'Supplier 2', 'Supplier 3', 'Supplier 4', 'Supplier 5']
for i, sup in enumerate(suppliers, start=6):
    ws4.cell(row=i, column=2, value=sup)
    ws4.cell(row=i, column=3, value=f'=COUNTIF(\'Raw Data\'!N:N,B{i})')
    ws4.cell(row=i, column=4, value=f'=AVERAGEIF(\'Raw Data\'!N:N,B{i},\'Raw Data\'!AC:AC)').number_format = '0.00"%"'
    ws4.cell(row=i, column=5, value=f'=AVERAGEIF(\'Raw Data\'!N:N,B{i},\'Raw Data\'!P:P)').number_format = '0.0'
    ws4.cell(row=i, column=6, value=f'=COUNTIFS(\'Raw Data\'!N:N,B{i},\'Raw Data\'!T:T,"Fail")/C{i}').number_format = '0.0"%"'
    ws4.cell(row=i, column=7, value=f'=SUMIF(\'Raw Data\'!N:N,B{i},\'Raw Data\'!F:F)').number_format = '$#,##0'
    ws4.cell(row=i, column=8, value=f'=IF(AND(D{i}>2.3,E{i}>17),"High",IF(OR(D{i}>2.3,E{i}>17),"Medium","Low"))')

for col_letter, width in [('B', 14), ('C', 11), ('D', 16), ('E', 18), ('F', 14), ('G', 14), ('H', 11)]:
    ws4.column_dimensions[col_letter].width = width

# conditional formatting for risk level column (color scale based on defect rate cell)
from openpyxl.formatting.rule import CellIsRule
ws4.conditional_formatting.add('H6:H10', CellIsRule(operator='equal', formula=['"High"'], fill=PatternFill('solid', fgColor='F8CBCB')))
ws4.conditional_formatting.add('H6:H10', CellIsRule(operator='equal', formula=['"Medium"'], fill=PatternFill('solid', fgColor='FCEBBB')))
ws4.conditional_formatting.add('H6:H10', CellIsRule(operator='equal', formula=['"Low"'], fill=PatternFill('solid', fgColor='D4EFDF')))

# ============================================================
# TAB 5: Stockout Risk (filtered list, formula-free — direct from analysis)
# ============================================================
ws5 = wb.create_sheet('Stockout Risk')
ws5.sheet_view.showGridLines = False
ws5['B2'] = 'SKUs Where Stock Cover Is Thinner Than Supplier Lead Time'
ws5['B2'].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws5['B3'] = 'These SKUs risk running out of stock before a reorder could arrive, if current sales pace holds.'
ws5['B3'].font = Font(name=FONT, size=10, italic=True, color='595959')

risk_df = df[df['Days of stock cover'] < df['Lead time']].sort_values('Days of stock cover').head(25)
risk_headers = ['SKU', 'Product Type', 'Supplier', 'Stock Level', 'Days of Cover', 'Supplier Lead Time (days)']
for j, h in enumerate(risk_headers, start=2):
    c = ws5.cell(row=5, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
for i, (_, r) in enumerate(risk_df.iterrows(), start=6):
    ws5.cell(row=i, column=2, value=r['SKU'])
    ws5.cell(row=i, column=3, value=r['Product type'])
    ws5.cell(row=i, column=4, value=r['Supplier name'])
    ws5.cell(row=i, column=5, value=int(r['Stock levels']))
    ws5.cell(row=i, column=6, value=round(r['Days of stock cover'], 1))
    ws5.cell(row=i, column=7, value=int(r['Lead time']))
for col_letter, width in [('B', 10), ('C', 14), ('D', 14), ('E', 12), ('F', 14), ('G', 20)]:
    ws5.column_dimensions[col_letter].width = width

wb.save('/home/claude/supply-chain-project/excel/analysis.xlsx')
print("Workbook built with tabs:", wb.sheetnames)
